# import win32com.client as win32                                                         # Using xlwings, save the Excel file as a PDF
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
import openpyxl
from PIL import Image  
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
# import xlsxwriter
from mpl_toolkits.mplot3d import Axes3D
import os
import glob
import datetime
import random
import string
import subprocess
# import pythoncom
# pythoncom.CoInitialize()
# %matplotlib widget


def Code_dissagregation(df, LAT, LON, file_csv_name, folder_path, project_name):
    all_means = []
    
    workbook = openpyxl.Workbook()  # Create an Excel workbook
    
    # df['%'] = abs(df['rlz46'] * 100)                # Create a new column '%' that contains the absolute values of 'rlz67' multiplied by 100
    rlz_column = [col for col in df.columns if col.startswith('rlz')][0]
    df['%'] = abs(df[rlz_column] * 100)  # Create a new column '%' that contains the absolute values of 'rlz' multiplied by 100
    # df['%'] = abs(df['rlz28'] * 100)                # Create a new column '%' that contains the absolute values of 'rlz28' multiplied by 100
    df['Return Period'] = abs(df['poe'])                      # Create a new column '%' that contains the absolute values of 'poe'
    
    # Create a new DataFrame that contains the sum of the contribution for each combination of magnitude, distance and type of fault contribution
    df_grouped = df.groupby(['imt','poe','mag', 'dist', 'trt','Return Period'])['%'].sum().reset_index()

    # Create a different color for each type of fault contribution using the 'Pastel1' colormap, which provides a sequence of pastel colors.
    trt_colors = {trt: color for trt, color in zip(df_grouped['trt'].unique(), plt.cm.Pastel1(np.linspace(-0, 1, len(df_grouped['trt'].unique())*2)))}

    Accel =  df_grouped['imt'].unique()
    df_grouped_total = df_grouped
    
    # Display IMT values found in the uploaded .csv file in a row format using a centered table
    st.markdown('**IMT Values:**')
    st.markdown("In the **.csv** file you uploaded, the following Intensity Measure Types (IMT) are found:")
    col1, col2, col3 = st.columns([0.3,1,0.3])
    with col1:
        st.metric(label= "",value="")
    with col2:
        imts_file = pd.DataFrame(Accel).transpose()
        st.write(imts_file)
    with col3:
        st.metric(label= "",value="")
    
    # Display TRT values found in the uploaded .csv file in a row format using a centered table
    st.markdown('**TRT Values:**')
    st.markdown("In the **.csv** file you uploaded, the following Seismic Return Periods (TR) are found for each value of IMT:")
    col1, col2, col3 = st.columns([0.60,1,0.60])
    with col1:
        st.metric(label= "",value="")
    with col2:
        values_POE = df_grouped['Return Period'].unique()
        values_RETURT = np.around((1/(1-(1 - values_POE)**(1/50))))
        dataF_values_RETURN = pd.DataFrame(values_RETURT).transpose()
        st.write(dataF_values_RETURN)
    with col3:
        st.metric(label= "",value="")

    
    # num_imts = len(Accel)                                                           # number of selected IMT's [nomalie use the all  IMT's contained in "Accel"]
    num_imts = 1                                                                      # number of selected IMT's [nomalie use the all  IMT's contained in "Accel"]

    images = {}                                                                         # Initialize a dictionary to store images

    # Initialize the progress bar
    progress_bar = st.progress(0)
    progress_text = st.empty()

    # Calculate the total number of iterations for the progress bar
    total_iterations = num_imts * len(df_grouped_total['Return Period'].unique())
    current_iteration = 0
    
    
    # Dictionary mapping CSV faults to PEER faults
    fault_mapping = {
        'Active Shallow Crust': ['Strike Slip (SS)', 'Normal/Oblique', 'Reverse/Oblique', 'SS+Normal', 'SS+Reverse', 'Normal+Reverse'],
        'Stable Shallow Crust': ['Normal/Oblique', 'Reverse/Oblique'],
        'Subduction Interface': ['Reverse/Oblique'],
        'Subduction IntraSlab': ['Normal/Oblique']
    }  

    df_grouped_filtered_group = pd.DataFrame()
    
    # Loop through each selected IMT
    for zz in range(num_imts):   
        IMT_selected = zz
        prueba = df_grouped_total[df_grouped_total['imt'] == Accel[IMT_selected]]
        df_grouped = prueba

        RETU = df_grouped['Return Period'].unique()
        RETURT = np.around((1/(1-(1 - RETU)**(1/50))))
        
        TRT_Rmeans_Mmeans = pd.DataFrame(columns=['TRT','Rmean','Mmean'])
        TRT_Rmeans_Mmeans_Faults = pd.DataFrame(columns=['TRT', 'Rmean', 'Mmean', 'Dominant_Fault'])
        
        # Loop through each return period
        for ReturT in range(len(RETURT)):
            # Filter the DataFrame to include only the rows where 'Return Period' is equal to each value of 
            df_grouped_filtered = df_grouped[df_grouped['Return Period'] == RETU[ReturT]]
            
            # Get the minimum and maximum percentage values for each type of fault contribution
            trt_min_max = df_grouped_filtered.groupby('trt')['%'].agg(['min', 'max']).reset_index()
            
            # Merge the minimum and maximum values with the original grouped DataFrame
            df_grouped_filtered = pd.merge(df_grouped_filtered, trt_min_max, on='trt')
            
            Rmean = (df_grouped_filtered["dist"] * df_grouped_filtered["%"]).sum() / df_grouped_filtered["%"].sum()     # Estimation of avarage distance Rmean
            Mmean = (df_grouped_filtered["mag"] * df_grouped_filtered["%"]).sum() / df_grouped_filtered["%"].sum()      # Estimation of avarage Magnitud Rmean

            df_grouped_filtered_group = pd.concat([df_grouped_filtered_group, df_grouped_filtered], ignore_index=True)
            TRT_Rmeans_Mmeans.loc[ReturT] = [RETURT[ReturT], Rmean, Mmean]                                              # Saving the Data [Rmean, Mmean]

            
            # Find the fault type with the maximum contribution for each return period
            max_contrib_trt = df_grouped_filtered.loc[df_grouped_filtered['%'].idxmax(), 'trt']
            # Assign the PEER equivalence according to the dominant fault type
            peer_faults = fault_mapping.get(max_contrib_trt, ['Unknown'])
            # Create a string with the corresponding PEER faults
            peer_faults_str = ', '.join(peer_faults)

            # Add a new row to the DataFrame with the dominant fault information and its PEER equivalence
            TRT_Rmeans_Mmeans_Faults = TRT_Rmeans_Mmeans_Faults.append({
                'TRT': RETURT[ReturT],
                'Rmean': Rmean,
                'Mmean': Mmean,
                'Dominant_Fault': max_contrib_trt,
                'PEER_Equivalent_Faults': peer_faults_str
            }, ignore_index=True)
            
            
            # Plot the disaggregation data        
            fig = plt.figure(figsize=(16, 8))
            fig.patch.set_facecolor('white')
            ax = fig.add_subplot(111, projection='3d')

            stacked_bars = {}
            
            z_max = max(df_grouped_filtered[df_grouped_filtered['poe'] == RETU[ReturT]]['%'])           # Calculted max value of % acumulated contribution for each POE
            
            if z_max > 0:
                z_max = z_max 
            else:
                z_max = 1
            

            for i, row in df_grouped_filtered.iterrows():
                x = row['dist']  
                y = row['mag']  
                z = row['%'] / z_max                                                                    # Normalized value of % acumulated contribution for each POE
                color = trt_colors[row['trt']]

                if (x, y) in stacked_bars:
                    z_bottom = stacked_bars[(x, y)]
                    stacked_bars[(x, y)] += z
                else:
                    z_bottom = 0
                    stacked_bars[(x, y)] = z

                barX_width = 5
                barY_width = 0.05
                ax.bar3d(x-(barX_width/2), y-(barY_width/2), z_bottom, barX_width, barY_width, z, color=color, edgecolor='black', linewidth=0.1, shade=True)

            legend_elements = []
            for trt, color in trt_colors.items():
                if trt in df_grouped_filtered['trt'].unique():
                    min_value = df_grouped_filtered.loc[df_grouped_filtered['trt'] == trt, 'min'].values[0]
                    max_value = df_grouped_filtered.loc[df_grouped_filtered['trt'] == trt, 'max'].values[0]
                    label = f'{trt} (Min: {min_value:.2f}, Max: {max_value:.2f})'
                    legend_elements.append(mpatches.Patch(color=color, label=label))

            ax.legend(handles=legend_elements, title='Contribution of Geological Fault Type', bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0., frameon=False)
            ax.grid(which='both', axis='x', alpha=0.5)                                                                                                          # Add gridlines to both the x and y axis
            ax.text2D(0.85, 0.99, '© TorreFuerte', transform=ax.transAxes, color=(0, 0, 1), alpha=0.5,
                fontsize=10, verticalalignment='top', horizontalalignment='left')
            ax.set_xlabel('R [km]')
            ax.set_ylabel('Mw')
            ax.set_zlabel('Hazard Contribution [%]')
            plt.title(f'Disaggregation of Seismic Hazard, TR = {RETURT[ReturT]:.0f} years, Project Name: {project_name}\nIntensity of Measure: {Accel[IMT_selected]} , [LAT = {LAT:.4f} -- LON = {LON:.4f}]\n--- Rmean = {Rmean:.2f} km ---------------- Mmean = {Mmean:.2f} Mw ---\nPowered by SDTE © TorreFuerte\nFile Selected {file_csv_name}', fontsize=11, color=(0, 0, 1))

            ax.get_proj = lambda: np.dot(Axes3D.get_proj(ax), np.diag([1.3, 0.9, 1, 1]))
            ax.view_init(elev=20, azim=-64)
            ax.xaxis.pane.set_edgecolor('black')
            ax.yaxis.pane.set_edgecolor('black')
            ax.zaxis.pane.set_edgecolor('black')
            ax.grid(True)

                
            ############ Save this plot ################
            at2_file_name = Accel[zz]  # Get the name of each IMT selected
            image_path = os.path.join(folder_path, at2_file_name + '_' + str(int(RETURT[ReturT])) + '.png')
            fig.savefig(image_path)

            # Load the saved image
            images["image" + str(ReturT)] = Image.open(image_path)
            
  
            # st.pyplot(fig)
            
            # Update progress
            current_iteration += 1
            progress_percentage = current_iteration / total_iterations
            progress_bar.progress(progress_percentage)
            progress_text.text(f"Processing: {progress_percentage:.1%} complete")
        
        # Safe of Means of each IMT
        title_df = pd.DataFrame({Accel[IMT_selected]}, index=[-1])
        TRT_Rmeans_Mmeans_IMT = pd.concat([title_df, TRT_Rmeans_Mmeans_Faults]).reset_index(drop=True)
        all_means.append(TRT_Rmeans_Mmeans_IMT)                                                 # Append the DataFrame to the list
        
        
    
        # Plot Rmean for each IMT's
        fig2, ax2 = plt.subplots(figsize=(16/1.5, 9/1.5))
        colors = plt.cm.winter(np.linspace(0, 1, len(TRT_Rmeans_Mmeans['TRT'])))
        handles = []                                                                            # variable for saving list of values for each 'bar'
        # Draw individual 'bar's'
        for trt, rmean, color in zip(TRT_Rmeans_Mmeans['TRT'], TRT_Rmeans_Mmeans['Rmean'], colors):
            bar = ax2.bar(trt, rmean, 30, edgecolor='black', linewidth=0.7, color=color, alpha=0.7)
            handles.append(bar[0])
            ax2.text(trt, rmean, f'R_JB = {rmean:.2f} [Km]', ha='center', va='bottom', fontsize=6)
        plt.title(f'R_JB(km): Distance from the epicenter to the fault, Project Name: {project_name}\nPowered by SDTE © TorreFuerte\nIntensity of Measure: {Accel[IMT_selected]} , [LAT = {LAT:.4f} -- LON = {LON:.4f}]', fontsize=11, color=(0, 0, 1))
        ax2.set_xlabel('TRs(years)')
        ax2.set_ylabel('R_JB [km]')
        ax2.text(0.85, 0.99, '© TorreFuerte', transform=ax2.transAxes, color=(0, 0, 1), alpha=0.5,
                    fontsize=10, verticalalignment='top', horizontalalignment='left')
        ax2.legend(handles=handles, labels=TRT_Rmeans_Mmeans['TRT'].tolist(), title='TRT [years]' , bbox_to_anchor=(1, 1), loc='upper left', borderaxespad=0., frameon=False)
        ############ Save this plot ################
        image_path = os.path.join(folder_path, at2_file_name + '_Rmean' + '.png')
        fig2.savefig(image_path)
        # Load the saved image
        images["image" + str(ReturT+1)] = Image.open(image_path) 
        
        
        # Plot Mmean for each IMT's
        fig3, ax3 = plt.subplots(figsize=(16/1.5, 9/1.5))
        colors = plt.cm.cividis(np.linspace(0, 1, len(TRT_Rmeans_Mmeans['TRT'])))
        handles = []                                                                            # variable for saving list of values for each 'bar'
        # Draw individual 'bar's'
        for trt, mmean, color in zip(TRT_Rmeans_Mmeans['TRT'], TRT_Rmeans_Mmeans['Mmean'], colors):
            bar = ax3.bar(trt, mmean, 30, edgecolor='black', linewidth=0.7, color=color, alpha=0.7)
            handles.append(bar[0])
            ax3.text(trt, mmean, f'M_mean = {mmean:.2f} [Mw]', ha='center', va='bottom', fontsize=6)
        plt.title(f'M_mean(Mw): Seismic Magnitud, Project Name: {project_name}\nPowered by SDTE © TorreFuerte\nIntensity of Measure: {Accel[IMT_selected]} , [LAT = {LAT:.4f} -- LON = {LON:.4f}]', fontsize=11, color=(0, 0, 1))
        ax3.set_xlabel('IMT [TRs(years)]')
        ax3.set_ylabel('Mmean [Mw]')
        ax3.text(0.85, 0.99, '© TorreFuerte', transform=ax3.transAxes, color=(0, 0, 1), alpha=0.5,
                    fontsize=10, verticalalignment='top', horizontalalignment='left')
        ax3.legend(handles=handles, labels=TRT_Rmeans_Mmeans['TRT'].tolist(), title='TRT [years]' , bbox_to_anchor=(1, 1), loc='upper left', borderaxespad=0., frameon=False)
        ############ Save this plot ################
        image_path = os.path.join(folder_path, at2_file_name + '_Mmean' + '.png')
        fig3.savefig(image_path)
        # Load the saved image
        images["image" + str(ReturT+2)] = Image.open(image_path)
        
        
        

        ####################################################################### Excel File Processing and Customization ###################################################################################
     
        worksheet = workbook.create_sheet()                                                         # Create a new sheet for the current IMT selected
        worksheet.title = f"{at2_file_name}"                                                        # Set the sheet name using the IMT selected.
        # Customization of the Current Excel Sheet
        worksheet.page_setup.scale = 48
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.26
        worksheet.page_margins.bottom = 0.42
        worksheet.page_setup.horizontalCentered = True
        worksheet.page_setup.verticalCentered = True
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE                            # Set orientation to landscape

        # Establecer el título para que se repita en todas las páginas al imprimir
        worksheet.oddHeader.left.text = at2_file_name
        worksheet.evenHeader.left.text = at2_file_name
        
        disclaimer_text = "Disclaimer: The data in this file is provided for academic purposes only. The user assumes full responsibility for the use and application of this data./ Powered by SDTE © TorreFuerte"
        worksheet.oddFooter.center.text = disclaimer_text
        worksheet.evenFooter.center.text = disclaimer_text
        
        
        for hh in range(len(RETURT) + 2):
            img_key = "image" + str(hh)
            if img_key in images:
                img = openpyxl.drawing.image.Image(images[img_key].filename)  # Convertir a objeto de imagen de openpyxl
                worksheet.add_image(img, 'B' + str(1 + hh * 40))

    
    file_excel_name = "Disaggregation.xlsx"
    file_excel_path = os.path.join(folder_path, file_excel_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    workbook.save(file_excel_path)
    file_pdf_name = "Disaggregation.pdf"                                            
    excel_to_pdf_path = os.path.join(folder_path, file_pdf_name)                     
    
    # Waiting time for excel file to open and transform in PDF file
    import time
    time.sleep(5)
    excel = win32.DispatchEx("Excel.Application")
    wb = excel.Workbooks.Open(os.path.abspath(os.path.join(folder_path, file_excel_name)))
    wb.ExportAsFixedFormat(0, excel_to_pdf_path)
    wb.Close(False)
    excel.Quit()


    # Progrres bar
    progress_bar.progress(1.0)
    st.success("All calculations were executed successfully.")
    
    # Display a success message with the folder path
    st.markdown('##### :sparkles: **Results Ready!**')
    st.markdown('* Use this directory to see all the results:')
    st.success(folder_path)
    st.markdown('**Sample Result:**')
    st.markdown('* Mean Values of **Rjb** and **Mw** for Each **IMT**')
    
       
    # Display Table's of Means of each IMT
    col1, col2, col3 = st.columns([0.08,1,0.08])
    with col1:
        st.metric(label= "",value="")
    with col2:
        TRT_Rmeans_Mmeans_IMTs = pd.concat(all_means, ignore_index=True)
        st.write(TRT_Rmeans_Mmeans_IMTs)
    with col3:
        st.metric(label= "",value="")
    
    
    
    ################################################# Save the DataFrame to a new Excel file without creating the workbook first #############################################
    file_excel_name2 = "Summary_Results_Rmean_Mmean.xlsx"
    file_excel_path2 = os.path.join(folder_path, file_excel_name2)

    # Save the DataFrame to the created Excel worksheet
    with pd.ExcelWriter(file_excel_path2, engine='openpyxl') as writer:
        # Access the workbook and sheet
        workbook = writer.book
        worksheet = workbook.create_sheet("Summary_Results_Rmean_Mmean")

        # Add disclaimer in the first row
        disclaimer_text = "Disclaimer: The data in this file is provided for academic purposes only. The user assumes full responsibility for the use and application of this data./ Powered by SDTE © TorreFuerte"
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TRT_Rmeans_Mmeans_IMTs.columns))
        worksheet.cell(row=1, column=1).value = disclaimer_text

        # Write DataFrame to Excel, starting from the second row
        TRT_Rmeans_Mmeans_IMTs.to_excel(writer, sheet_name="Summary_Results_Rmean_Mmean", startrow=1, index=False)

        # Adjust the column width to fit the content, starting from the second row
        for column_cells in worksheet.iter_cols(min_row=2, max_col=len(TRT_Rmeans_Mmeans_IMTs.columns)):
            length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length

        # Customization of the Current Excel Sheet
        worksheet.page_setup.scale = 84
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.26
        worksheet.page_margins.bottom = 0.42
        worksheet.page_setup.horizontalCentered = True
        worksheet.page_setup.verticalCentered = False
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE  # Set orientation to landscape
        
        disclaimer_text = "Disclaimer: The data in this file is provided for academic purposes only. The user assumes full responsibility for the use and application of this data./ Powered by SDTE © TorreFuerte"
        worksheet.oddFooter.center.text = disclaimer_text
        worksheet.evenFooter.center.text = disclaimer_text


    # file_pdf_name2 = "Summary_Results_Rmean_Mmean.pdf"                                            
    # excel2_to_pdf2_path = os.path.join(folder_path, file_pdf_name2)                     

    # # Waiting time for excel file to open and transform in PDF file
    # import time
    # time.sleep(5)
    # excel = win32.DispatchEx("Excel.Application")
    # wb = excel.Workbooks.Open(os.path.abspath(os.path.join(folder_path, file_excel_name2)))
    # wb.ExportAsFixedFormat(0, excel2_to_pdf2_path)
    # wb.Close(False)
    # excel.Quit()

    
    
    
    
    ################################################# Save the DataFrame to a new Excel file without creating the workbook first #############################################
    file_excel_name3 = "MetaData.xlsx"
    file_excel_path3 = os.path.join(folder_path, file_excel_name3)
    
    # Save the DataFrame to the created Excel worksheet
    with pd.ExcelWriter(file_excel_path3, engine='openpyxl') as writer:
        df_grouped_filtered_group.to_excel(writer, sheet_name="MetaData", index=False)
        # Access the workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets["MetaData"]

        # Adjust the column width to fit the content
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length

        # Customization of the Current Excel Sheet
        worksheet.page_setup.scale = 130
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.26
        worksheet.page_margins.bottom = 0.42
        worksheet.page_setup.horizontalCentered = True
        worksheet.page_setup.verticalCentered = False
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE                            # Set orientation to landscape

    # file_pdf_name3 = "MetaData.pdf"                                            
    # excel3_to_pdf3_path = os.path.join(folder_path, file_pdf_name3)                     

    # # Waiting time for excel file to open and transform in PDF file
    # import time
    # time.sleep(5)
    # excel = win32.DispatchEx("Excel.Application")
    # wb = excel.Workbooks.Open(os.path.abspath(os.path.join(folder_path, file_excel_name3)))
    # wb.ExportAsFixedFormat(0, excel3_to_pdf3_path)
    # wb.Close(False)
    # excel.Quit()

       
    ############################################# Display one of the saved images #############################################
    if images:
        st.markdown('**Sample Result:**')
        st.image(images["image0"], caption='Sample Disaggregation Plot')
        st.image(images["image4"], caption='Sample Disaggregation Plot')
        st.image(images["image5"], caption='Sample Disaggregation Plot')

    
    return TRT_Rmeans_Mmeans_IMT 
